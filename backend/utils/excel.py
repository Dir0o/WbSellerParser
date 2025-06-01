from typing import List, Dict, Any
from datetime import datetime, time
import openpyxl.utils.cell
from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from schemas.wb import SellerOut

def generate_excel_search(data: list, filename: str = "search_results.xlsx") -> str:
    """
    Генерация Excel-файла для результатов поиска SellerDetail.
    Форматирование такое же, как в generate_excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchResults"

    if not data:
        wb.save(filename)
        return filename

    ws.append(
        [
            "Айди продавца",
            "Наименование магазина",
            "Ссылка",
            "ИНН",
            "ОГРН",
            "ОГРНИП",
            "Регистратор",
            "Продаж",
            "Дата создание магазина",
            "Телефон",
            "Телефон 2",
            "Телефон 3",
            "Телефон 4",
            "Телефон 5",
            "Телефон 6",
            "Телефон 7",
            "Телефон 8",
            "Телефон 9",
            "Телефон 10",
            "Почта"
        ]
    )
    MAX_PHONES = 10
    for seller in data:
        d = [
            seller.seller_id,
            seller.store_name,
            seller.url,
            seller.inn,
            seller.ogrn if seller.ogrn else None,
            seller.ogrnip if seller.ogrnip else None,
            seller.tax_office,
            seller.saleCount if seller.saleCount is not None else None,
            str(seller.reg_date).split("+")[0] if seller.reg_date else None,
        ]

        phones = seller.phone or []
        for i in range(MAX_PHONES):
            d.append(phones[i] if i < len(phones) else None)

        for email in (seller.email or []):
            d.append(email)
        ws.append(
            d
        )

    # Форматирование заголовка
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(size=14)
        cell.fill = PatternFill("solid", fgColor="bfbfbf")

    # Границы для всех ячеек
    thin = Side(border_style="thin", color="16365C")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, bottom=thin)

    ws.column_dimensions["A"].width = 16.67
    ws.column_dimensions["B"].width = 40.67
    ws.column_dimensions["C"].width = 44.67
    ws.column_dimensions["D"].width = 15.67
    ws.column_dimensions["E"].width = 17.67
    ws.column_dimensions["F"].width = 19.67
    ws.column_dimensions["G"].width = 60.67
    ws.column_dimensions["H"].width = 15.67
    ws.column_dimensions["I"].width = 24.67
    ws.column_dimensions["J"].width = 20.67
    ws.column_dimensions["K"].width = 20.67
    ws.column_dimensions["L"].width = 20.67
    ws.column_dimensions["M"].width = 20.67
    ws.column_dimensions["N"].width = 20.67
    ws.column_dimensions["O"].width = 20.67
    ws.column_dimensions["P"].width = 20.67
    ws.column_dimensions["Q"].width = 20.67
    ws.column_dimensions["R"].width = 20.67
    ws.column_dimensions["S"].width = 20.67
    ws.column_dimensions["T"].width = 25.67

    ws.row_dimensions[1].height = 36.67

    wb.save(filename)
    return filename

def generate_excel(data: List[SellerOut], filename: str = "sellers.xlsx") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sellers"

    normalized: List[SellerOut] = []
    for item in data:
        if isinstance(item, dict):
            normalized.append(SellerOut(**item))
        else:
            normalized.append(item)
    data = normalized

    if not data:
        wb.save(filename)
        return filename

    ws.append(
        [
            "Айди продавца",
            "Наименование магазина",
            "Ссылка",
            "ИНН",
            "ОГРН",
            "ОГРНИП",
            "Регистратор",
            "Продаж",
            "Дата создание магазина",
            "Телефон",
            "Телефон 2",
            "Телефон 3",
            "Телефон 4",
            "Телефон 5",
            "Телефон 6",
            "Телефон 7",
            "Телефон 8",
            "Телефон 9",
            "Телефон 10",
            "Почта"
        ]
    )
    MAX_PHONES = 10
    for seller in data:
        d = [
                seller.seller_id,
                seller.store_name,
                seller.url,
                seller.inn,
                seller.ogrn if seller.ogrn else None,
                seller.ogrnip if seller.ogrnip else None,
                seller.tax_office,
                seller.saleCount,
                str(seller.reg_date).split('+')[0] if seller.reg_date else None,
            ]
        phones = seller.phone or []
        for i in range(MAX_PHONES):
            d.append(phones[i] if i < len(phones) else None)

        for email in (seller.email or []):
            d.append(email)
        ws.append(
            d
        )

    for row in ws.iter_rows(max_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=14)
            cell.fill = PatternFill("solid", fgColor="bfbfbf")

    thin_border = Side(border_style="thin", color="16365C")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin_border, right=thin_border, bottom=thin_border)

    ws.column_dimensions["A"].width = 16.67
    ws.column_dimensions["B"].width = 40.67
    ws.column_dimensions["C"].width = 44.67
    ws.column_dimensions["D"].width = 15.67
    ws.column_dimensions["E"].width = 17.67
    ws.column_dimensions["F"].width = 19.67
    ws.column_dimensions["G"].width = 60.67
    ws.column_dimensions["H"].width = 15.67
    ws.column_dimensions["I"].width = 24.67
    ws.column_dimensions["J"].width = 20.67
    ws.column_dimensions["K"].width = 20.67
    ws.column_dimensions["L"].width = 20.67
    ws.column_dimensions["M"].width = 20.67
    ws.column_dimensions["N"].width = 20.67
    ws.column_dimensions["O"].width = 20.67
    ws.column_dimensions["P"].width = 20.67
    ws.column_dimensions["Q"].width = 20.67
    ws.column_dimensions["R"].width = 20.67
    ws.column_dimensions["S"].width = 20.67
    ws.column_dimensions["T"].width = 25.67

    ws.row_dimensions[1].height = 36 + 0.67

    wb.save(filename)
    return filename